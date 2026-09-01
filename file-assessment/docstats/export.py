"""Export the scan result to Excel / CSV."""

from __future__ import annotations

import csv
import os

from .analyzer import ScanResult, TableInfo

FILE_HEADERS = [
    "File",
    "Type",
    "Path",
    "Pages",
    "Tables",
    "Top-level tables",
    "Figures",
    "Small figures dropped",
    "Complex tables",
    "Split across pages",
    "Tables with nested tables",
    "Nested tables",
    "Tables with merged cells",
    "Split without header",
    "Total cells",
    "Cells merged away",
    "Highest score",
    "Error",
]

TABLE_HEADERS = [
    "File",
    "Type",
    "Table",
    "Level",
    "Page",
    "Pages spanned",
    "Rows",
    "Columns",
    "Rows×Columns",
    "Cells (score)",
    "Cells merged away",
    "Complex",
    "Horizontal merges",
    "Vertical merges",
    "Has merge",
    "Nested tables",
    "Is nested",
    "Split",
    "Break after row",
    "Repeat header",
    "Issues",
    "Caption",
    "Preview",
    "Path",
]

FIGURE_HEADERS = [
    "File",
    "Figure #",
    "Kind",
    "Page",
    "Width×Height (cm)",
    "% of page area",
    "Counted",
    "In table",
    "Path",
]


def page_text(t: TableInfo) -> str:
    if t.page_start is None:
        return "?"
    base = (
        str(t.page_start)
        if t.page_start == t.page_end
        else f"{t.page_start}-{t.page_end}"
    )
    return f"~{base}" if t.page_is_approx else base


def file_rows(result: ScanResult) -> list[list]:
    rows = []
    for d in result.docs:
        rows.append(
            [
                d.name,
                d.file_type,
                d.path,
                d.pages or "",
                d.n_tables_all,
                d.n_tables_top,
                d.n_figures,
                d.n_figures_small,
                d.n_complex,
                d.n_split,
                d.n_nested_parent,
                d.n_nested_child,
                d.n_merged,
                d.n_split_no_header,
                d.total_cells,
                d.total_merged_cells,
                d.max_score,
                d.error,
            ]
        )
    return rows


def table_rows(result: ScanResult) -> list[list]:
    rows = []
    for doc in result.docs:
        for t in doc.tables:
            rows.append(_table_row(t, doc.file_type))
    return rows


def _table_row(t: TableInfo, file_type: str) -> list:
    return [
        t.file_name,
        file_type,
        t.label,
        t.level,
        page_text(t),
        t.pages_spanned or "",
        t.rows,
        t.grid_cols,
        t.grid_cells,
        t.cells,
        t.merged_cells,
        "x" if t.is_complex else "",
        t.merged_h,
        t.merged_v,
        "x" if t.has_merge else "",
        t.nested_direct,
        "x" if t.is_nested else "",
        "x" if t.is_split else "",
        ", ".join(str(r) for r in t.split_rows),
        "x" if t.repeat_header else "",
        "; ".join(t.issue_tags()),
        t.caption,
        t.preview,
        t.file,
    ]


def figure_rows(result: ScanResult) -> list[list]:
    """List every figure, dropped ones included, with a column saying which count."""
    return [
        [
            f.file_name,
            f.index,
            f.kind,
            f.page or "",
            f.size_text,
            round(f.area_percent, 1) if f.size_known else "",
            "x" if f.counted else "",
            "x" if f.in_table else "",
            f.file,
        ]
        for f in result.all_figures()
    ]


def export_csv(result: ScanResult, folder: str) -> list[str]:
    os.makedirs(folder, exist_ok=True)
    written = []
    for name, headers, rows in (
        ("files.csv", FILE_HEADERS, file_rows(result)),
        ("tables.csv", TABLE_HEADERS, table_rows(result)),
        ("figures.csv", FIGURE_HEADERS, figure_rows(result)),
    ):
        path = os.path.join(folder, name)
        with open(path, "w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.writer(fh)
            writer.writerow(headers)
            writer.writerows(rows)
        written.append(path)
    return written


def export_xlsx(result: ScanResult, path: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    totals = result.totals()

    ws = wb.active
    ws.title = "Summary"
    ws.append(["WORD / PDF DOCUMENT STATISTICS"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["Folder", result.root])
    ws.append(["Complex-table threshold", f"≥ {result.complex_threshold} cells"])
    ws.append(
        [
            "Figure area threshold",
            f"≥ {result.figure_min_area * 100:.0f}% of the page"
            if result.figure_min_area > 0
            else "count every figure (logos included)",
        ]
    )
    ws.append([])
    summary = [
        ("Files", totals["files"]),
        ("  Word files", totals["files_word"]),
        ("  PDF files", totals["files_pdf"]),
        ("Files with errors", totals["files_error"]),
        ("Total pages", totals["pages"]),
        ("Total tables", totals["tables"]),
        ("  of which top-level", totals["tables_top"]),
        ("Total figures", totals["figures"]),
        ("  small figures dropped", totals["figures_small"]),
        ("Complex tables", totals["complex"]),
        ("Tables split across pages", totals["split"]),
        ("Tables containing nested tables", totals["nested_parent"]),
        ("Tables nested inside another", totals["nested_child"]),
        ("Tables with merged cells", totals["merged"]),
        ("Split without repeating the header", totals["split_no_header"]),
        ("Total cells", totals["cells"]),
        ("Total cells merged away", totals["merged_cells"]),
    ]
    for label, value in summary:
        ws.append([label, value])
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 70

    head_fill = PatternFill("solid", fgColor="1F4E79")
    head_font = Font(bold=True, color="FFFFFF")

    for title, headers, rows in (
        ("By file", FILE_HEADERS, file_rows(result)),
        ("Tables", TABLE_HEADERS, table_rows(result)),
        ("Figures", FIGURE_HEADERS, figure_rows(result)),
    ):
        sheet = wb.create_sheet(title)
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in rows:
            sheet.append(row)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for idx, header in enumerate(headers, start=1):
            width = max(10, min(46, len(header) + 4))
            if header in ("Preview", "Caption", "Path", "File"):
                width = 40
            sheet.column_dimensions[get_column_letter(idx)].width = width

    wb.save(path)
    return path
